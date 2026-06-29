"""Giữ deliverable cuối (.inp, .odb, BC-1 xydata) và ghi Excel tổng hợp."""

from __future__ import annotations

import tempfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from abaqus_postprocess import rf3_xydata_output_path
from file_io import remove_tree, replace_file, run_with_retry

SUMMARY_EXCEL_NAME = "TongHop_KetQua.xlsx"
BC1_NODE_SET = "BC-1"
ODB_OUTPUT_SUFFIX = "-TL"
EXCEL_HEADERS = ("Tên model", "Max RF3 (BC-1)")
EXCEL_SAVE_RETRIES = 15
EXCEL_SAVE_DELAY = 1.0


@dataclass
class ModelSummaryRow:
    model_name: str
    max_rf3_bc1: float


def summary_excel_path(output_root: Path) -> Path:
    return output_root.resolve() / SUMMARY_EXCEL_NAME


def deliverable_odb_path(output_dir: Path, job_name: str) -> Path:
    return output_dir.resolve() / f"{job_name}{ODB_OUTPUT_SUFFIX}.odb"


def deliverable_keep_paths(
    output_dir: Path,
    *,
    job_name: str,
    imperfection_inp: Path,
) -> set[Path]:
    output_dir = output_dir.resolve()
    return {
        imperfection_inp.resolve(),
        deliverable_odb_path(output_dir, job_name).resolve(),
        rf3_xydata_output_path(output_dir, job_name, BC1_NODE_SET).resolve(),
    }


def parse_xydata_max_rf3(path: Path) -> float:
    """Lấy giá trị Y lớn nhất trong file BC-1 xydata."""
    if not path.is_file():
        raise FileNotFoundError(f"Không tìm thấy xydata: {path}")
    y_max: float | None = None
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        text = line.strip()
        if not text or text.lower().startswith("x"):
            continue
        parts = text.replace(",", ".").split()
        if len(parts) < 2:
            continue
        try:
            y_val = float(parts[1])
        except ValueError:
            continue
        if y_max is None or y_val > y_max:
            y_max = y_val
    if y_max is None:
        raise ValueError(f"Không đọc được dữ liệu Y trong {path.name}")
    return y_max


def cleanup_model_output_dir(output_dir: Path, *, keep_paths: set[Path]) -> list[str]:
    """Xóa mọi file/thư mục trong output_dir trừ các deliverable giữ lại."""
    output_dir = output_dir.resolve()
    keep = {p.resolve() for p in keep_paths}
    removed: list[str] = []

    if not output_dir.is_dir():
        return removed

    for item in sorted(output_dir.iterdir(), key=lambda p: p.name.lower()):
        resolved = item.resolve()
        if resolved in keep:
            continue
        try:
            if item.is_dir():
                remove_tree(item)
            else:
                item.unlink()
            removed.append(item.name)
        except OSError:
            pass
    return removed


def _migrate_legacy_time_column(ws: Worksheet) -> None:
    """File cũ 3 cột (có Thời gian chạy) -> 2 cột."""
    header_b = str(ws.cell(1, 2).value or "").strip().lower()
    if "thời gian" not in header_b and "thoi gian" not in header_b:
        return
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 2).value = ws.cell(row_idx, 3).value
    ws.delete_cols(2, 1)
    ws.cell(1, 1, EXCEL_HEADERS[0])
    ws.cell(1, 2, EXCEL_HEADERS[1])


def _ensure_headers(ws: Worksheet) -> None:
    if ws.max_row < 1 or not ws.cell(1, 1).value:
        ws.cell(1, 1, EXCEL_HEADERS[0])
        ws.cell(1, 2, EXCEL_HEADERS[1])
        return
    _migrate_legacy_time_column(ws)
    ws.cell(1, 1, EXCEL_HEADERS[0])
    ws.cell(1, 2, EXCEL_HEADERS[1])


def _find_model_row(ws: Worksheet, model_name: str) -> int | None:
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row_idx, 1).value
        if cell_val is not None and str(cell_val).strip() == model_name:
            return row_idx
    return None


def _upsert_row_in_sheet(ws: Worksheet, row: ModelSummaryRow) -> None:
    _ensure_headers(ws)
    target = _find_model_row(ws, row.model_name)
    if target is None:
        target = ws.max_row + 1
    ws.cell(target, 1, row.model_name)
    ws.cell(target, 2, row.max_rf3_bc1)


def _save_workbook_atomic(wb: openpyxl.Workbook, excel_path: Path) -> None:
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        suffix=".xlsx",
        dir=excel_path.parent,
        delete=False,
    ) as tmp:
        tmp_path = Path(tmp.name)
    try:
        wb.save(tmp_path)
        replace_file(tmp_path, excel_path)
    finally:
        wb.close()
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def upsert_summary_row(excel_path: Path, row: ModelSummaryRow) -> Path:
    """Thêm hoặc cập nhật một dòng — giữ nguyên các dòng khác, ghi qua file tạm."""
    excel_path = excel_path.resolve()

    def do_upsert() -> None:
        if excel_path.is_file():
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Ket qua"
        _upsert_row_in_sheet(ws, row)
        _save_workbook_atomic(wb, excel_path)

    run_with_retry(
        do_upsert,
        excel_path,
        retries=EXCEL_SAVE_RETRIES,
        delay=EXCEL_SAVE_DELAY,
    )
    return excel_path


def write_summary_excel(excel_path: Path, rows: list[ModelSummaryRow]) -> Path:
    """Ghi nhiều dòng — merge theo tên model."""
    excel_path = excel_path.resolve()
    for row in rows:
        upsert_summary_row(excel_path, row)
    return excel_path


def finalize_model_deliverables(
    output_dir: Path,
    *,
    job_name: str,
    imperfection_inp: Path,
    model_name: str | None = None,
    on_progress=None,
) -> tuple[Path, float, list[str]]:
    """
    Đọc max RF3 từ BC-1 xydata, dọn thư mục, cập nhật Excel tổng hợp.
    Trả về (bc1_xydata_path, max_rf3, removed_files).
    """
    output_dir = output_dir.resolve()
    bc1_path = rf3_xydata_output_path(output_dir, job_name, BC1_NODE_SET)
    max_rf3 = parse_xydata_max_rf3(bc1_path)

    keep = deliverable_keep_paths(
        output_dir,
        job_name=job_name,
        imperfection_inp=imperfection_inp,
    )
    removed = cleanup_model_output_dir(output_dir, keep_paths=keep)

    label = model_name or imperfection_inp.name
    excel_path = summary_excel_path(output_dir.parent)
    upsert_summary_row(
        excel_path,
        ModelSummaryRow(model_name=label, max_rf3_bc1=max_rf3),
    )

    if on_progress:
        on_progress(f"Bước 3: Max RF3 (BC-1) = {max_rf3:g}")
        on_progress(f"Đã dọn {len(removed)} file/thư mục — giữ .inp, .odb, BC-1 xydata")
        on_progress(f"Excel tổng hợp → {excel_path.name}")

    return bc1_path, max_rf3, removed


__all__ = [
    "BC1_NODE_SET",
    "ModelSummaryRow",
    "SUMMARY_EXCEL_NAME",
    "cleanup_model_output_dir",
    "deliverable_keep_paths",
    "finalize_model_deliverables",
    "parse_xydata_max_rf3",
    "summary_excel_path",
    "upsert_summary_row",
    "write_summary_excel",
]
