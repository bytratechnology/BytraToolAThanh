import math
import tempfile
from pathlib import Path

import openpyxl

from file_io import locked_file_message, replace_file, run_with_retry, write_text
from inp_parser import parse_nodes_from_inp
from matlab_writer import write_matlab_file
from paths import DEFAULT_PATHS, ProjectPaths

SHEET_SECTION_A = "SectionA"
START_ROW = 3
TOL = 1e-3
COORD_OFFSET = 7.5


def write_nodes_to_section_a(nodes, excel_template, excel_output, start_row=3):
    """
    Đọc Excel mẫu, dán tọa độ node vào SectionA (H=x, I=y, J=z), ghi ra output.
    File mẫu gốc không bị thay đổi.
    """
    excel_template = Path(excel_template)
    excel_output = Path(excel_output)
    excel_output.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(excel_template)
    try:
        ws = wb[SHEET_SECTION_A]

        for i, (_, x, y, z) in enumerate(nodes):
            row = start_row + i
            ws[f"H{row}"] = x
            ws[f"I{row}"] = y
            ws[f"J{row}"] = z

        with tempfile.NamedTemporaryFile(
            suffix=".xlsx",
            dir=excel_output.parent,
            delete=False,
        ) as tmp:
            tmp_path = Path(tmp.name)

        try:
            def save_excel():
                wb.save(tmp_path)
                replace_file(tmp_path, excel_output)

            run_with_retry(save_excel, excel_output)
        except PermissionError:
            raise
        except OSError as exc:
            if _is_file_locked_oserror(exc):
                raise PermissionError(locked_file_message(excel_output)) from exc
            raise
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except OSError:
                    pass
    finally:
        wb.close()


def _is_file_locked_oserror(exc: OSError) -> bool:
    return getattr(exc, "winerror", None) == 32 or "being used by another process" in str(exc)


def _format_number(value):
    if value == int(value):
        return str(int(value))
    return str(value)


def write_matrix_txt(nodes, matrix_output):
    """Ghi tọa độ node (x, y, z) vào Matrix.txt, mỗi dòng cách nhau bằng tab."""
    lines = [
        f"{_format_number(x)}\t{_format_number(y)}\t{_format_number(z)}"
        for _, x, y, z in nodes
    ]
    write_text(matrix_output, "\n".join(lines) + "\n")


def _dist(point):
    return math.hypot(point[0], point[1])


def _farthest_in_quadrant(xy_points, x_positive, y_positive):
    """Điểm xa gốc tọa độ nhất trong một góc phần tư."""
    candidates = [
        p
        for p in xy_points
        if (p[0] > 0 if x_positive else p[0] < 0)
        and (p[1] > 0 if y_positive else p[1] < 0)
    ]
    if not candidates:
        quadrant = (
            f"x{'>' if x_positive else '<'}0, y{'>' if y_positive else '<'}0"
        )
        raise ValueError(f"Không tìm thấy điểm trong góc phần tư ({quadrant})")
    return max(candidates, key=_dist)


def find_lip_midpoint_between(point_a, point_b, nodes, lip_x=None):
    """Điểm giữa trên môi (X23/Y23, X78/Y78): cùng x môi, nằm giữa 2 điểm, gần trục hoành nhất."""
    x_lip = lip_x if lip_x is not None else point_a[0]
    y_low = min(point_a[1], point_b[1])
    y_high = max(point_a[1], point_b[1])

    xy_points = list({(x, y) for _, x, y, _ in nodes})
    on_lip = [
        p
        for p in xy_points
        if abs(p[0] - x_lip) < TOL and y_low < p[1] < y_high
    ]
    if not on_lip:
        raise ValueError(f"Không tìm thấy điểm giữa trên môi x={x_lip}")

    return min(on_lip, key=lambda p: abs(p[1]))


def select_11_points(nodes):
    """
    Chọn 11 tọa độ (x, y) theo thứ tự MATLAB:
    - Điểm 1-4, 6-9: từ điểm xa nhất mỗi góc phần tư, offset ±7.5 theo trục
    - Điểm 5: gần trục tung nhất
    - Điểm 23, 78: giữa môi giữa điểm 2-3 và 7-8
    """
    xy_points = list({(x, y) for _, x, y, _ in nodes})
    offset = COORD_OFFSET

    # A: Q2 (x<0, y>0), B: Q3, C: Q4, D: Q1
    point_a = _farthest_in_quadrant(xy_points, x_positive=False, y_positive=True)
    point_b = _farthest_in_quadrant(xy_points, x_positive=False, y_positive=False)
    point_c = _farthest_in_quadrant(xy_points, x_positive=True, y_positive=False)
    point_d = _farthest_in_quadrant(xy_points, x_positive=True, y_positive=True)

    p1 = (point_a[0] + offset, point_a[1])
    p2 = (point_a[0], point_a[1] - offset)
    p3 = (point_b[0], point_b[1] + offset)
    p4 = (point_b[0] + offset, point_b[1])
    p6 = (point_c[0] - offset, point_c[1])
    p7 = (point_c[0], point_c[1] + offset)
    p8 = (point_d[0], point_d[1] - offset)
    p9 = (point_d[0] - offset, point_d[1])

    near_y_axis = min(xy_points, key=lambda p: (abs(p[0]), -_dist(p)))

    min_x = min(p[0] for p in xy_points)
    max_x = max(p[0] for p in xy_points)

    mid_23 = find_lip_midpoint_between(p2, p3, nodes, lip_x=min_x)
    mid_78 = find_lip_midpoint_between(p7, p8, nodes, lip_x=max_x)

    return [
        ("Điểm 1 (X1,Y1) - Q2, A + offset x", p1),
        ("Điểm 2 (X2,Y2) - Q2, A - offset y", p2),
        ("Điểm 23 (X23,Y23) - giữa điểm 2 và 3", mid_23),
        ("Điểm 3 (X3,Y3) - Q3, B + offset y", p3),
        ("Điểm 4 (X4,Y4) - Q3, B + offset x", p4),
        ("Điểm 5 (X5,Y5) - trục tung", near_y_axis),
        ("Điểm 6 (X6,Y6) - Q4, C - offset x", p6),
        ("Điểm 7 (X7,Y7) - Q4, C + offset y", p7),
        ("Điểm 78 (X78,Y78) - giữa điểm 7 và 8", mid_78),
        ("Điểm 8 (X8,Y8) - Q1, D - offset y", p8),
        ("Điểm 9 (X9,Y9) - Q1, D - offset x", p9),
    ]


def write_selected_points(selected_points, output_path):
    """Ghi 11 điểm đo ra file text."""
    lines = [
        "# 11 tọa độ đo (X, Y) chọn từ Matrix",
        "# STT\tMô tả\tX\tY",
    ]
    for i, (label, (x, y)) in enumerate(selected_points, start=1):
        lines.append(f"{i}\t{label}\t{_format_number(x)}\t{_format_number(y)}")
    write_text(output_path, "\n".join(lines) + "\n")


def _notify(on_progress, message: str):
    print(message, flush=True)
    if on_progress:
        on_progress(message)


def run_processing(paths: ProjectPaths | None = None, on_progress=None):
    """Bước 1: đọc .inp, ghi Excel/Matrix, chọn 11 tọa độ, tạo file MATLAB."""
    paths = (paths or DEFAULT_PATHS).resolve()

    _notify(on_progress, "Bước 1: Kiểm tra file nguồn...")
    paths.validate_sources()
    paths.ensure_work_dir()

    _notify(
        on_progress,
        f"Bước 1: Đang đọc node từ {paths.inp_source.name} (*Node đầu → *Element)...",
    )
    nodes = parse_nodes_from_inp(paths.inp_source)

    _notify(on_progress, f"Bước 1: Đã đọc {len(nodes)} node. Đang ghi Excel...")
    write_nodes_to_section_a(
        nodes,
        paths.excel_template,
        str(paths.excel_output),
        START_ROW,
    )
    _notify(on_progress, f"Bước 1: Đã ghi Excel → {paths.excel_output.name}")

    _notify(on_progress, "Bước 1: Đang ghi Matrix.txt...")
    write_matrix_txt(nodes, str(paths.matrix_output))
    _notify(on_progress, f"Bước 1: Đã ghi Matrix → {paths.matrix_output.name}")

    _notify(on_progress, "Bước 1: Đang chọn 11 tọa độ đo...")
    selected_points = select_11_points(nodes)
    write_selected_points(selected_points, str(paths.selected_points_output))
    _notify(on_progress, f"Bước 1: Đã ghi 11 tọa độ → {paths.selected_points_output.name}")

    _notify(on_progress, "Bước 1: Đang tạo file MATLAB...")
    write_matlab_file(
        selected_points,
        matlab_template=str(paths.matlab_template),
        matlab_output=str(paths.matlab_output),
        node_count=len(nodes),
    )
    _notify(on_progress, f"Bước 1: Đã ghi MATLAB → {paths.matlab_output.name}")
    _notify(
        on_progress,
        f"Bước 1: Hoàn tất. Chạy Bước 2 để tạo {paths.inp_result.name}",
    )

    return len(nodes)


def run_pipeline():
    """Mở GUI; xử lý chỉ chạy khi người dùng bấm nút."""
    from gui import launch_gui

    print("Mở giao diện...")
    launch_gui()


if __name__ == "__main__":
    run_pipeline()
